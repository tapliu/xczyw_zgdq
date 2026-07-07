import os
import json
import openpyxl

_DATA_DIR = os.path.join(os.path.dirname(__file__), '..', '..', '..', 'server', 'src', 'data')
_XLSX_PATH = os.path.join(_DATA_DIR, 'relation_network.xlsx')
_JSON_CACHE = os.path.join(_DATA_DIR, 'relations.json')


def _load_data():
    if os.path.isfile(_JSON_CACHE):
        with open(_JSON_CACHE, 'r', encoding='utf-8') as f:
            raw = json.load(f)
        raw_cf = raw.get('char_factions', {})
        char_factions = {}
        for n, info in raw_cf.items():
            if isinstance(info, dict):
                primary = info.get('primary', '群雄')
                all_f = info.get('all', [primary])
            else:
                primary = info
                all_f = [primary]
            char_factions[n] = {'primary': primary, 'all': all_f}
        factions = raw.get('factions', {})
        char_relations = raw.get('char_relations', {})
        faction_chars = {}
        for n, info in char_factions.items():
            faction_chars.setdefault(info['primary'], []).append(n)
        if '群雄' not in factions:
            factions['群雄'] = {'lord': '', 'retainers': [], 'allies': [], 'enemies': []}
        return {
            'char_factions': char_factions,
            'faction_chars': faction_chars,
            'factions': factions,
            'char_relations': char_relations,
        }
    return None


def _load_xlsx():
    wb = openpyxl.load_workbook(_XLSX_PATH)
    return wb


def _clean_name(s):
    if not s:
        return ''
    s = s.strip()
    for ch in '（）()':
        s = s.replace(ch, '')
    return s


def _parse_name_list(s):
    if not s or s == '无':
        return []
    parts = s.replace('、', ',').replace('，', ',').split(',')
    result = []
    for p in parts:
        p = _clean_name(p)
        if p:
            result.append(p)
    return result


def build_faction_data():
    wb = _load_xlsx()

    ws1 = wb['sheet1']
    char_factions = {}
    faction_chars = {}
    for r in range(2, ws1.max_row + 1):
        gid = ws1.cell(r, 1).value
        name = ws1.cell(r, 2).value
        faction = ws1.cell(r, 4).value
        if name and faction:
            name = _clean_name(name)
            factions = [_clean_name(f) for f in faction.replace('→', ',').split(',')]
            primary = factions[0]
            char_factions[name] = {'id': gid, 'primary': primary, 'all': factions}
            faction_chars.setdefault(primary, []).append(name)
            for f in factions[1:]:
                faction_chars.setdefault(f, []).append(name)

    ws2 = wb['sheet2']
    factions = {}
    seen = {}
    for r in range(2, ws2.max_row + 1):
        name = ws2.cell(r, 1).value
        if not name:
            continue
        name = _clean_name(name)
        lord = _clean_name(ws2.cell(r, 2).value or '')
        retainers_str = ws2.cell(r, 3).value or ''
        allies_str = ws2.cell(r, 4).value or ''
        enemies_str = ws2.cell(r, 5).value or ''

        retainers = _parse_name_list(retainers_str)
        allies = _parse_name_list(allies_str)
        enemies = _parse_name_list(enemies_str)

        if name in seen:
            existing = seen[name]
            for ret in retainers:
                if ret not in existing['retainers']:
                    existing['retainers'].append(ret)
            for ally in allies:
                if ally not in existing['allies']:
                    existing['allies'].append(ally)
            for enemy in enemies:
                if enemy not in existing['enemies']:
                    existing['enemies'].append(enemy)
            if lord and not existing['lord']:
                existing['lord'] = lord
        else:
            entry = {'lord': lord, 'retainers': retainers, 'allies': allies, 'enemies': enemies}
            seen[name] = entry
            factions[name] = entry

    for name, data in seen.items():
        factions[name] = data

    # Merge single-character factions into 群雄
    from collections import Counter
    fc = Counter()
    for name, info in char_factions.items():
        fc[info['primary']] += 1
    single_factions = {f for f, c in fc.items() if c == 1}
    for name, info in list(char_factions.items()):
        if info['primary'] in single_factions:
            info['primary'] = '群雄'
            info['all'] = ['群雄']
    for sf in single_factions:
        faction_chars.pop(sf, None)
        factions.pop(sf, None)
    faction_chars['群雄'] = [n for n, info in char_factions.items() if info['primary'] == '群雄']
    if '群雄' not in factions:
        factions['群雄'] = {'lord': '', 'retainers': [], 'allies': [], 'enemies': []}

    ws3 = wb['sheet3']
    char_relations = {}
    seen_rels = set()
    for r in range(2, ws3.max_row + 1):
        a = _clean_name(ws3.cell(r, 2).value or '')
        rel = _clean_name(ws3.cell(r, 3).value or '')
        b = _clean_name(ws3.cell(r, 4).value or '')
        note = ws3.cell(r, 5).value or ''
        if not a or not b or not rel:
            continue
        key = (a, rel, b)
        if key in seen_rels:
            continue
        seen_rels.add(key)
        char_relations.setdefault(a, []).append({'type': rel, 'target': b, 'note': note})

    # Write JSON cache for fast loading next time
    try:
        cache = {
            'char_factions': {n: {'primary': info['primary'], 'all': info['all']} for n, info in char_factions.items()},
            'factions': {f: {'lord': info['lord'], 'allies': info['allies'], 'enemies': info['enemies']} for f, info in factions.items()},
            'char_relations': char_relations,
        }
        with open(_JSON_CACHE, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    return {
        'char_factions': char_factions,
        'faction_chars': faction_chars,
        'factions': factions,
        'char_relations': char_relations,
    }


DATA = None


def get_data():
    global DATA
    if DATA is None:
        DATA = _load_data()
        if DATA is None:
            DATA = build_faction_data()
    return DATA


def get_faction(char_name):
    d = get_data()
    info = d['char_factions'].get(char_name)
    if info:
        return info['primary']
    return '群雄'


def get_factions(char_name):
    d = get_data()
    info = d['char_factions'].get(char_name)
    if info:
        return info['all']
    return ['群雄']


def get_lord(faction_name):
    d = get_data()
    f = d['factions'].get(faction_name)
    if f:
        return f['lord']
    return None


def is_hostile(faction_a, faction_b):
    if not faction_a or not faction_b:
        return False
    d = get_data()
    fa = d['factions'].get(faction_a)
    if fa and faction_b in fa['enemies']:
        return True
    fb = d['factions'].get(faction_b)
    if fb and faction_a in fb['enemies']:
        return True
    return False


def is_ally(faction_a, faction_b):
    if not faction_a or not faction_b:
        return False
    if faction_a == faction_b:
        return True
    d = get_data()
    fa = d['factions'].get(faction_a)
    if fa and faction_b in fa['allies']:
        return True
    fb = d['factions'].get(faction_b)
    if fb and faction_a in fb['allies']:
        return True
    return False


def get_char_relations(char_name):
    d = get_data()
    return d['char_relations'].get(char_name, [])


def get_all_factions():
    d = get_data()
    return list(d['factions'].keys())
